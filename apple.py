from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
import time
import pickle
import re
import openpyxl
from openpyxl.utils import get_column_letter
import telebot
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook, Workbook

bot = telebot.TeleBot('6922962992:AAEFLaHCfs1lU6vdkQzWh-GucKilbeEZFmg')

piastr = 678484052
TARGET_CHAT_ID = '678484052'
MESSAGE_THREAD_ID = 'General'
#        4090    4070ti   4080
prices = [178000, 68000, 90000]
prices4090MM = ['280000', '42']

dict = {
    "iphone 11 64": 36800,
    "iphone 11 128": 43800,
    "iphone 12 64": 40500,
    "iphone 12 128": 45500,
    "iphone 12 256": 51500,
    "iphone 12 mini 256": 43500,
    "iphone 13 128": 52000,
    "iphone 13 256": 70000,
    "iphone 13 mini 128": 56700,
    "iphone 13 mini 256": 62000,
    "iphone 13 mini 512": 67000,
    "iphone 14 128": 57000,
    "iphone 14 256": 73400,
    "iphone 14 512": 87500,
    "iphone 14 plus 128": 68000,
    "iphone 14 plus 256": 85000,
    "iphone 14 plus 512": 87000,
    "iphone 14 pro 128": 86000,
    "iphone 14 pro 256": 101500,
    "iphone 14 pro 512": 112500,
    "iphone 14 pro 1024": 114700,
    "iphone 14 pro max 128": 98500,
    "iphone 14 pro max 256": 98000,
    "iphone 14 pro max 512": 113000,
    "iphone 14 pro max 1024": 125700,
    "iphone 15 128": 70000,
    "iphone 15 256": 81500,
    "iphone 15 512": 92000,
    "iphone 15 plus 128": 77000,
    "iphone 15 plus 256": 86500,
    "iphone 15 plus 512": 108000,
    "iphone 15 pro 128": 93900,
    "iphone 15 pro 256": 105500,
    "iphone 15 pro 512": 127000,
    "iphone 15 pro 1024": 137900,
    "iphone 15 pro max 256": 115000,
    "iphone 15 pro max 512": 135500,
    "iphone 15 pro max 1024": 160800
}


def sort_list(lst):
    def get_number(item):
        # Преобразование строки в нижний регистр для удобства поиска
        item_lower = item[0].lower()
        # Разделение строки на слова и поиск числа после "iphone"
        words = item_lower.split()
        if "iphone" in words:
            index = words.index("iphone") + 1
            if index < len(words) and words[index].isdigit():
                return int(words[index])
        return 0

    # Сортировка списка по числу после "iphone" по убыванию
    sorted_list = sorted(lst, key=lambda x: -get_number(x))
    return sorted_list

    # Сортировка списка по числу после "iphone" по убыванию
    sorted_list = sorted(lst, key=lambda x: -get_number(x[0]))
    return sorted_list


def get_products_apple(bot, timesleep, url, page):
    bot = bot

    options = Options()
    options.add_argument('--headless')
    options.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36")
    options.add_argument("--disable-blink-features=AutomationControlled")
    driver = webdriver.Chrome(options=options)
    driver.maximize_window()
    driver.get('https://megamarket.ru/shop/gadzhet-stor/#?page=1')
    time.sleep(2)
    """cookies = pickle.load(open("cookies", "rb"))
    for cookie in cookies:
        driver.add_cookie(cookie)"""

    time.sleep(2)
    driver.refresh()
    time.sleep(2)

    while True:
        print('Apple работаю')
        products = []
        xx = url.replace('~', '1')
        driver.get(xx)
        time.sleep(1)
        try:
            cc = driver.find_element(By.CLASS_NAME, 'filter-footer').get_attribute('textContent')
            cc = re.sub(r'\D', '', cc)
            print(cc)
            num_pages = int(cc) // 24
            #num_pages = 5

            for i in range(num_pages):
                print(i)
                try:
                    driver.get(url.replace('~', f'{i + 1}'))
                    time.sleep(0.3)
                except:
                    break

                x = driver.find_elements(By.CLASS_NAME, 'item-block')
                xxx = driver.find_elements(By.CLASS_NAME, 'ddl_product_link')
                result = [xxx[i] for i in range(len(xxx)) if i % 2 == 1]
                for i in range(len(result)):
                    element = []


                    element.append(result[i].accessible_name)
                    try:
                        rub = x[i].find_element(By.CLASS_NAME, 'item-price').find_element(By.CSS_SELECTOR,
                                                                                          'span').get_attribute(
                            'innerHTML')[:-7]
                        rub = re.sub(r'\W', '', rub)
                        element.append(int(rub))
                        sber_spasibo = x[i].find_element(By.CLASS_NAME, 'item-bonus').find_element(By.CLASS_NAME,
                                                                                                   'bonus-amount').get_attribute(
                            'innerHTML')
                        sber_spasibo = re.sub(r'\W', '', sber_spasibo)
                        element.append(int(sber_spasibo))
                        percent = (int(sber_spasibo) / int(rub)) * 100
                        element.append("{:.0f}".format(percent))
                        element.append(result[i].get_attribute('href'))
                        product_id = x[i].find_element(By.CSS_SELECTOR, 'a').get_attribute('data-product-id')
                        element.append(product_id)
                        products.append(element)
                    except:
                        continue
            print(len(products))
            list_x = []

            products = sort_list(products)



            """ name = str(result[i].accessible_name)
                name = name.lower()
                name = name[name.find('iphone'):]
                if name.find('гб'):
                    name = name[:name.find('гб')]
                if name.find('gb'):
                    name = name[:name.find('gb')]
                name = name.rstrip()
                name = re.sub(r'[^a-zA-Z0-9\s]', '', name)
                print(name, dict[name])"""


            if len(products) != 0:
                for i in products:
                    try:
                        name = str(i[0])
                        name = name.lower()
                        name = name[name.find('iphone'):]
                        if name.find('гб'):
                            name = name[:name.find('гб')]
                        if name.find('gb'):
                            name = name[:name.find('gb')]
                        name = name.rstrip()
                        name = re.sub(r'[^a-zA-Z0-9\s]', '', name)
                        price_sell = dict[name]
                        win = i[2] * 0.85 - (i[1] - price_sell)
                        if int("{:.0f}".format((int(win) / (int(i[1]) - int(price_sell)) * 100))) > 70 and int("{:.0f}".format(int(win))) > 30000:
                            elem = []
                            elem.append(i[0])
                            elem.append("{:.0f}".format(int(win)))
                            elem.append(i[1] - price_sell)
                            elem.append("{:.0f}".format((int(win) / (int(i[1]) - int(price_sell)) * 100)))
                            elem.append(i[1])
                            elem.append(price_sell)
                            elem.append(i[2])
                            elem.append(i[3])
                            elem.append(i[4])
                            list_x.append(elem)
                            # ["Название", "Цена покупки", 'Выгода', "Цена продажи", "Бонусов", "Процент", "Ссылка"]

                    except Exception as e:
                        print(e)
                        print(f'Нет такого - {i[0]}')
                        continue

                if len(list_x) != 0:
                    sorted_list = sorted(list_x, key=lambda x: int(x[3]), reverse=True)

                    result = []
                    seen = set()

                    for sublist in sorted_list:
                        # Преобразуем подсписок в кортеж
                        tuple_sublist = tuple(sublist)

                        # Проверяем, видели ли мы этот подсписок ранее
                        if tuple_sublist not in seen:
                            result.append(sublist)
                            seen.add(tuple_sublist)

                    try:
                        # Попытка загрузить существующий файл
                        workbook = load_workbook('apple.xlsx')
                    except FileNotFoundError:
                        # Если файл не найден, создаем новый
                        workbook = Workbook()
                    sheet_names = workbook.sheetnames
                    sheet_active = workbook[sheet_names[page]]
                    """# Определение имени листа в зависимости от переданного числа
                    if page == 1:
                        sheet = 'Айфоны'
                        sheet_active = workbook[sheet]
                    elif page == 2:
                        sheet = 'Наушники'
                        sheet_active = workbook[sheet]
                    elif page == 3:
                        sheet = 'Планшеты'
                        sheet_active = workbook[sheet]"""

                    # Получение или создание листа

                    # Подписываем колонки


                    for row_index, row in enumerate(result, start=1):
                        for col_index, value in enumerate(row, start=1):
                            sheet_active.cell(row=row_index, column=col_index).value = value

                        cell = sheet_active.cell(row=row_index, column=len(row))
                        cell.value = row[-1]
                        cell.hyperlink = row[-1]  # Присваиваем гиперссылку ячейке

                    for col in sheet_active.columns:
                        max_length = max(len(str(cell.value)) for cell in col if
                                         cell.value)  # Находим самую длинную строку в столбце
                        adjusted_width = (
                                                     max_length + 2) * 1.2  # Увеличиваем ширину столбца на 20% от максимальной длины
                        sheet_active.column_dimensions[get_column_letter(col[0].column)].width = adjusted_width
                    columns = ['Название', "Выгода", "Заморожено", " Проц. соотнош.", "Цена покуп.", "Цена прод.", "Бонус", "Процент", "Ссылка"]
                    #columns = ["Название", "Цена покупки", 'Выгода', "Цена продажи", "Бонусов", "Процент", "Ссылка"]
                    for col_num, value in enumerate(columns, 1):
                        sheet_active.cell(row=1, column=col_num, value=value)
                    # Сохраняем файл
                    workbook.save("apple.xlsx")
                    with open("apple.xlsx", 'rb') as f:
                        #, disable_notification=True
                        bot.send_document(TARGET_CHAT_ID, f, disable_notification=True)
            time.sleep(timesleep)
        except Exception as e:
            print(f'Error: {e}')
            continue


if __name__ == '__main__':
    get_products_apple(bot, 4000,
                       'https://megamarket.ru/catalog/smartfony-apple/page-~/#?filters=%7B"4CB2C27EAAFC4EB39378C4B7487E6C9E"%3A%5B"1"%5D%7D',
                       0)
